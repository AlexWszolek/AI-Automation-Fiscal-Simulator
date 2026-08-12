"""지역별고용조사 (Local Area Labour Force Survey): 시도 × occupation-major employment →
tidy CSV for the descriptive exposure map (workstream D — no provincial fiscal claims).

Source: 국가데이터처 (Statistics Korea) press release 「2025년 상반기 지역별고용조사
취업자의 산업 및 직업별 특성」 (2025-10-28), attachment
"('17년~'25년) … 통계표(참고용).xlsx", sheet 「4. 시도 직업별 취업자」.
Board page: kostat.go.kr/board.es?act=view&bid=211&list_no=439006
Attachment:  kostat.go.kr/boardDownload.es?bid=211&list_no=439006&seq=5

    .venv/bin/python scripts/fetch_korea_region_occupation.py [--parse-only]

Writes data/raw/korea/region_occupation.csv: period, region, occ_code (KSCO major 1-9),
emp_k (thousands, ALL employed — the survey covers self-employed too; the map discloses
that its occupation mix is all-employment, unlike the model's wage-worker cells).
"""
from __future__ import annotations

import argparse
import csv
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw" / "korea"
XLSX = RAW / "lafs_region_occupation_2017_2025.xlsx"
OUT = RAW / "region_occupation.tidy.csv"
LF_OUT = RAW / "region_labour_force.tidy.csv"
URL = "https://kostat.go.kr/boardDownload.es?bid=211&list_no=439006&seq=5"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")

REGIONS = ["서울특별시", "부산광역시", "대구광역시", "인천광역시", "광주광역시",
           "대전광역시", "울산광역시", "세종특별자치시", "경기도", "강원특별자치도",
           "충청북도", "충청남도", "전북특별자치도", "전라남도", "경상북도",
           "경상남도", "제주특별자치도"]


def fetch() -> None:
    subprocess.run(["curl", "-sL", "-A", UA, URL, "-o", str(XLSX)], check=True)
    assert XLSX.stat().st_size > 500_000, "download looks truncated"


def parse() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["4. 시도 직업별 취업자"]
    rows = list(ws.iter_rows(values_only=True))
    clean = lambda v: str(v).strip().replace("\xa0", "") if v is not None else ""
    periods = {clean(v): i for i, v in enumerate(rows[2]) if v and "20" in str(v)}

    out = []
    region = None
    for r in rows:
        a, b = clean(r[0]), clean(r[1])
        if a in REGIONS or a == "계":
            region = "전국" if a == "계" else a
        m = re.match(r"^(\d)\.", b)
        if region is None or not m:
            continue
        occ = int(m.group(1))
        for period, col in periods.items():
            v = r[col]
            if v in (None, "", "-"):
                continue
            out.append({"period": period, "region": region, "occ_code": occ,
                        "emp_k": float(v)})
    assert len(out) >= 18 * 9 * len(periods) * 0.95, f"parse looks truncated ({len(out)})"
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["period", "region", "occ_code", "emp_k"])
        w.writeheader()
        w.writerows(out)
    print(f"wrote {OUT} ({len(out)} rows, {len(periods)} periods)")

    # sheet 1: labour force / employed / unemployed per region-period (the denominator for
    # the unemployment-rate translation; 8 columns per period block)
    ws1 = wb["1. 시도 성별 경제활동인구 총괄"]
    rows1 = list(ws1.iter_rows(values_only=True))
    periods1 = {clean(v): i for i, v in enumerate(rows1[2]) if v and "20" in str(v)}
    lf = []
    region = None
    for r in rows1:
        a, b = clean(r[0]), clean(r[1])
        if a in REGIONS or a == "계":
            region = "전국" if a == "계" else a
        if region is None or b != "계":
            continue
        for period, col in periods1.items():
            vals = r[col:col + 4]                    # 15세이상, 경활, 취업자, 실업자
            if any(v in (None, "", "-") for v in vals):
                continue
            lf.append({"period": period, "region": region,
                       "labour_force_k": float(vals[1]), "employed_k": float(vals[2]),
                       "unemployed_k": float(vals[3])})
    assert len(lf) >= 18 * len(periods1) * 0.95, f"LF parse truncated ({len(lf)})"
    with open(LF_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["period", "region", "labour_force_k",
                                          "employed_k", "unemployed_k"])
        w.writeheader()
        w.writerows(lf)
    print(f"wrote {LF_OUT} ({len(lf)} rows)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--parse-only", action="store_true")
    a = ap.parse_args()
    if not a.parse_only:
        fetch()
    if not XLSX.exists():
        sys.exit("no xlsx — run without --parse-only to download")
    parse()
